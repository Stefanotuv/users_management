import openpyxl
from io import BytesIO
import pandas as pd
from django.http import HttpResponse


class FileManagement():

    # all utilities to modify the files

    # import excel file and create a correspondent json
    def excel_to_json(file_):
        # with no specified number of columns on the excel
        saved_entries = []
        not_saved_entries = []
        wb = openpyxl.load_workbook(file_)

        worksheet = wb[wb.sheetnames[0]]  # take the first sheet as the one with the data

        excel_data = list()
        excel_data_no_headers = list()

        # iterating over the rows and
        # getting value from each cell in row

        # table_name = model
        # model_class = getattr(sys.modules[__name__], table_name)

        i = 0  # i represent the rows
        headers_file = []
        data_dict = {}
        data_list = []
        data_dict_rows = {}
        list_dict_rows = []
        list_list_rows = []


        for row in worksheet.iter_rows():
            row_data = list()
            j = 0
            for cell in row:
                if cell.value:
                    row_data.append(str(cell.value))
                    if i != 0:
                        data_dict[headers_file[j]] = str(cell.value)
                        data_list.append(str(cell.value))
                else:
                    row_data.append(None)
                    if i != 0:
                        data_dict[headers_file[j]] = None
                        data_list.append(None)

                j = j + 1

            if i != 0:
                data_dict_rows['candidate_' + str(i)] = data_dict
                list_dict_rows.append(data_dict)
                list_list_rows.append(data_list)

            if i == 0:
                headers_file = row_data
            else:
                try:
                    excel_data_no_headers.append(row_data)
                    # model_class(**data_dict).save()save
                    print('saved question number:%s' % i)
                    saved_entries.append(i)
                except:
                    print('not saved')
                    not_saved_entries.append(i)
                    pass

            data_dict = {}
            data_list = []
            excel_data.append(row_data)
            i = i + 1
            json_file = {
                'file_name': file_.name,
                'headers':headers_file, #
                'candidates': list_dict_rows,
                'candidates_list': list_list_rows,
                'candidates_dictionary': data_dict_rows,
            }

        return json_file


    # manages the htmlk table posted from the screen
    def table_post_to_json(posted_items_tuple_list):
        headers = []
        items = []
        for tuple in posted_items_tuple_list:  # element is a tuple
            # if "header" in tuple[0]:
            if tuple[0].startswith("header"):
                # headers.append(element[7:(len(element))])
                headers.append(tuple[1])
                pass
            # elif "item" in tuple[0]:
            elif tuple[0].startswith("item"):
                items.append(tuple[1])
                pass
            else:
                pass

        jsondata = {}
        jsondata_candidate = {}
        candidates = []
        candidates_list = []
        list_for_candidates = []
        num_of_rows = int(len(items) / len(headers))
        index = 0
        for row_num in range(num_of_rows):

            for header in headers:
                jsondata_candidate[header] = items[index]
                list_for_candidates.append(items[index])
                index = index + 1

                pass
            jsondata["candidate_" + str(row_num + 1)] = jsondata_candidate
            candidates.append(jsondata_candidate)
            candidates_list.append(list_for_candidates)
            jsondata_candidate = {}
            list_for_candidates = []

        candidates_dictionary = jsondata
        return [headers, jsondata, candidates, candidates_list, candidates_dictionary]
        # return [headers,jsondata]

    def json_to_excel(json_document):
        with BytesIO() as b:
            df = pd.json_normalize(json_document['candidates'], max_level=5)
            # Use the StringIO object as the filehandle.
            writer = pd.ExcelWriter(b, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1')
            writer.save()
            # Set up the Http response.
            filename = 'django_simple.xlsx'
            response = HttpResponse(
                b.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
            return response
