from django.shortcuts import render

# Create your views here.
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, TemplateView, FormView, DetailView, UpdateView, CreateView
from .models import *
from .applications import ReviewConstraints, default_dimensions, dictionary_contraints, job_function, professional_category, company, job_title, mandatory_dimensions
from .applications import nationality, Calculations
from django.http import HttpResponse, HttpResponseRedirect, StreamingHttpResponse
from django.urls import reverse
import openpyxl

from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import EmailMessage
from users.models import User
from io import BytesIO
import json
from django.template.defaulttags import register


from django.shortcuts import get_object_or_404
import sys
# HomeView, CreateProjectView, ProjectView, ProjectListView

import pandas as pd

@method_decorator(login_required, name='dispatch')
class ProjectListView(ListView):
    model = Project
    context_object_name = 'viewlist'
    ordering = ['-last_modified']
    # ordering = ['-date_elaboration']
    # paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context

@method_decorator(login_required, name='dispatch')
class ProjectCreateView(CreateView):
    model = Project
    context_object_name = 'createproject'
    fields = '__all__'
    # ordering = ['-date_elaboration']
    # paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

@method_decorator(login_required, name='dispatch')
class ProjectDetailsView(DetailView):
    model = Project
    context_object_name = 'viewproject'

    # ordering = ['-date_elaboration']
    # paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

@method_decorator(login_required, name='dispatch')
class HomeView(TemplateView):
    model = Project
    context_object_name = 'viewproject'

    # ordering = ['-date_elaboration']
    # paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['created_projects'] = len(Project.objects.filter(status='created'))
        context['started_projects'] = len(Project.objects.filter(status='started'))
        context['reviewed_projects'] = len(Project.objects.filter(status='reviewed'))
        context['final_projects'] = len(Project.objects.filter(status='final'))

        return context

@method_decorator(login_required, name='dispatch')
class ProjectCreateFromFileView(CreateView):
    model = Project
    template_name = 'create.html'
    context_object_name = 'project_upload_from_file'
    fields = '__all__'


    # paginate_by = 10

    def get(self, request, *args, **kwargs):
        self.request.session['pk'] = kwargs['pk']

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pk'] =self.request.session['pk']
        project = Project.objects.get(pk=context['pk'])

        # if 'json_file' in context.keys(): context['json_file'] = project.json_document
        # if 'json_file_filename' in context.keys(): context['json_file_filename'] = project.json_document['file_name']
        # if 'json_file_headers' in context.keys(): context['json_file_headers'] = project.json_document['headers']
        # if 'json_file_candidates' in context.keys(): context['json_file_candidates'] = project.json_document['candidates']
        # if 'json_file_candidates_list' in context.keys(): context['json_file_candidates_list'] = project.json_document['candidates_list']
        # if 'json_file_candidates_dictionary' in context.keys(): context['json_file_candidates_dictionary'] = project.json_document['candidates_dictionary']
        # if 'project_name' in context.keys(): context['project_name'] = project.project_name

        context['json_file'] = project.json_document
        context['json_file_filename'] = project.json_document['file_name']
        context['json_file_headers'] = project.json_document['headers']
        context['json_file_candidates'] = project.json_document['candidates']
        context['json_file_candidates_list'] = project.json_document['candidates_list']
        context['json_file_candidates_dictionary'] = project.json_document['candidates_dictionary']
        context['project_name'] = project.project_name


        return context

    def post(self, request, *args, **kwargs):

        if (len(request.FILES)!=0) :

            excel_file = request.FILES["input_file"]
            json_file = excel_to_json(excel_file)

            # create the project
            new_project = Project(project_name=request.POST['project-name'],status='created',user=request.user.email,owner=request.user)
            new_project.save()

            # save original file
            temp_file = DocFile(file=excel_file)
            temp_file.save()
            # new_project.input_documents.add(temp_file)
            new_project.input_documents = temp_file

            # save json data
            new_project.json_document = json_file
            new_project.save()


            self.request.session['json_file'] = json_file
            self.request.session['json_file_filename'] = json_file['file_name']
            self.request.session['json_file_headers'] = json_file['headers']
            self.request.session['json_file_candidates'] = json_file['candidates']
            self.request.session['json_file_candidates_list'] = json_file['candidates_list']
            self.request.session['json_file_candidates_dictionary'] = json_file['candidates_dictionary']
            self.request.session['pk'] = new_project.pk


            # return HttpResponseRedirect(reverse('project_upload_from_file_view'), json_file)
            # return render(request,self.template_name)
            return super().post(request, *args, **kwargs)
        else:
            # test the post and see whether it is a submission or a review


            return super().get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectReviewView(UpdateView):
    model =Project
    template_name = 'review.html'
    context_object_name = 'reviewproject'
    fields ='__all__'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if (self.request.session['get_post'] == 'get'):
            json_file = context['object'].json_document

            context['get_post'] = self.request.session['get_post']
            context['pk'] = context['object'].pk
            context['json_file'] = json_file
            context['json_file_filename'] = json_file['file_name']
            context['json_file_headers'] = json_file['headers']
            context['json_file_candidates'] = json_file['candidates']
            context['json_file_candidates_list'] = json_file['candidates_list']
            context['json_file_candidates_dictionary'] = json_file['candidates_dictionary']
            # context['not_found_list'] = self.request.session['not_found_list']
            # context['check_value_columuns'] = self.request.session['check_value_columuns']

            # some old files do not have the issue table saved. for compactibility i added the if to exclude them
            if 'issues_table' in json_file.keys():
                context['issues_table'] = json_file['issues_table']
                context['table_headers'] = json_file['table_headers']
                context['dictionary_candidate_unmatched_fields'] = json_file['dictionary_candidate_unmatched_fields']
            else:
                context['issues_table'] = {}
                context['table_headers'] = {}
                context['dictionary_candidate_unmatched_fields'] = {}
            # context['candidate_with_issues_list'] = self.request.session['candidate_with_issues_list']

            project = Project.objects.get(pk=context['pk'])
            context['project_name'] = project.project_name
            context['message'] = ""


            nationality = [n.value for n in Nationality.objects.all()]
            job_title = [JT.value for JT in JobTitle.objects.all()]
            company = [C.value for C in Company.objects.all()]
            professional_category = [PC.value for PC in ProfessionalCategory.objects.all()]
            job_function = [JF.value for JF in JobFunction.objects.all()]

            context['nationality_list'] = nationality
            context['job_title'] = job_title
            context['company'] = company
            context['professional_category'] = professional_category
            context['job_function'] = job_function

        else:

            context['pk'] = context['object'].pk # to capture the pk of the project to be updated
            context['json_file'] =self.request.session['json_file']
            context['json_file_filename'] = self.request.session['json_file_filename']
            context['json_file_headers'] = self.request.session['json_file_headers']
            context['json_file_candidates'] = self.request.session['json_file_candidates']
            context['json_file_candidates_list'] = self.request.session['json_file_candidates_list']
            context['json_file_candidates_dictionary'] = self.request.session['json_file_candidates_dictionary']
            # context['not_found_list'] = self.request.session['not_found_list']
            # context['check_value_columuns'] = self.request.session['check_value_columuns']
            # context['candidate_with_issues_list'] = self.request.session['candidate_with_issues_list']
            project = Project.objects.get(pk=context['pk'])
            context['project_name'] = project.project_name
            # context['message'] = self.request.session['message']

            context['issues_table'] = project.json_document['issues_table']
            context['table_headers'] = project.json_document['table_headers']

            context['dictionary_candidate_unmatched_fields'] = self.request.session['dictionary_candidate_unmatched_fields']



            context['nationality_list'] = self.request.session['nationality_list']
            context['job_title'] = self.request.session['job_title']
            context['company'] = self.request.session['company']
            context['professional_category'] = self.request.session['professional_category']
            context['job_function'] = self.request.session['job_function']


        return context

    def get(self, request, *args, **kwargs):
        self.request.session['get_post'] = 'get'
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.request.session['get_post'] = 'post'
        # download as tuple the post items
        posted_items_tuple_list = list(request.POST.items())

        # the post contains the tuple for the key, value from the table posted for verification
        # we need to transform the post_items in a json format, however the post return a flat view for the items
        headers_json_table = table_post_to_json(posted_items_tuple_list)
        json_table = headers_json_table[1]
        headers = headers_json_table[0]
        candidates = headers_json_table[2]
        candidates_list = headers_json_table[3]
        candidates_dictionary = headers_json_table[4]
        review = ReviewConstraints()

        table_headers_issues_table_candidate_with_issues_list = review.issues_table(default_dimensions,headers,json_table,dictionary_contraints)
        table_headers = table_headers_issues_table_candidate_with_issues_list[0]
        issues_table = table_headers_issues_table_candidate_with_issues_list[1]
        dictionary_candidate_unmatched_fields = table_headers_issues_table_candidate_with_issues_list[2]

        project = Project.objects.get(pk=self.kwargs['pk'])
        project.status = 'started'
        project.json_document['issues_table'] = issues_table
        project.json_document['table_headers'] = table_headers
        project.json_document['dictionary_candidate_unmatched_fields'] = dictionary_candidate_unmatched_fields
        project.last_modified = datetime.now()

        project.save()

        self.request.session['issues_table'] = issues_table
        self.request.session['table_headers'] = table_headers
        self.request.session['json_file_candidates_dictionary'] = json_table
        self.request.session['dictionary_candidate_unmatched_fields'] = dictionary_candidate_unmatched_fields

        nationality = [n.value for n in Nationality.objects.all()]
        job_title = [JT.value for JT in JobTitle.objects.all()]
        company = [C.value for C in Company.objects.all()]
        professional_category = [PC.value for PC in ProfessionalCategory.objects.all()]
        job_function = [JF.value for JF in JobFunction.objects.all()]

        self.request.session['nationality_list'] = nationality
        self.request.session['job_title'] = job_title
        self.request.session['company'] = company
        self.request.session['professional_category'] = professional_category
        self.request.session['job_function'] = job_function

        self.request.session['message'] = ''



        return super().post(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectReviewedView(UpdateView):
    model =Project
    template_name = 'review.html'
    context_object_name = 'reviewproject'
    fields ='__all__'

    # def get_queryset(self):
    #
    #     self.obj = get_object_or_404(Project, pk=self.kwargs['project_name'])
    #     return Project.objects.filter(pk=self.obj.project_name)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if (self.request.session['get_post'] == 'get'):
            json_document = context['object'].json_document
            context['get_post'] = self.request.session['get_post']
            context['json_file'] = json_document
            context['json_file_filename'] = json_document['file_name']
            context['json_file_headers'] = json_document['headers']
            context['json_file_candidates'] = json_document['candidates']
            context['json_file_candidates_list'] = json_document['candidates_list']
            context['json_file_candidates_dictionary'] = json_document['candidates_dictionary']
            # context['issues_table'] = json_document['issues_table']
            # context['table_headers'] = json_document['table_headers']
            # context['dictionary_candidate_unmatched_fields'] = json_document[
            #     'dictionary_candidate_unmatched_fields']
            project = Project.objects.get(pk=context['object'].pk)
            context['project_name'] = project.project_name
            context['pk'] = context['object'].pk
        else:

            context['pk'] = self.request.session['pk']

            context['json_file'] =self.request.session['json_file']
            context['json_file_filename'] = self.request.session['json_file_filename']
            context['json_file_headers'] = self.request.session['json_file_headers']
            context['json_file_candidates'] = self.request.session['json_file_candidates']
            context['json_file_candidates_list'] = self.request.session['json_file_candidates_list']
            context['json_file_candidates_dictionary'] = self.request.session['json_file_candidates_dictionary']
            context['issues_table'] = self.request.session['issues_table']
            context['table_headers'] = self.request.session['table_headers']
            context['dictionary_candidate_unmatched_fields'] = self.request.session['dictionary_candidate_unmatched_fields']
            project = Project.objects.get(pk=context['pk'])
            context['project_name'] = project.project_name
        return context

    def get(self, request, *args, **kwargs):
        self.request.session['get_post'] = 'get'
        self.object = self.get_object()
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):



        self.request.session['get_post'] = 'post'
        posted_items_tuple_list = list(request.POST.items())
        self.request.session['posted_items_tuple_list'] = posted_items_tuple_list
        self.object = self.get_object()
        headers_json_table = table_post_to_json(posted_items_tuple_list)
        json_table = headers_json_table[1]
        headers = headers_json_table[0]
        candidates = headers_json_table[2]
        candidates_list = headers_json_table[3]
        candidates_dictionary = headers_json_table[4]
         # jsondata  candidates_dictionary are the same (to be reivewed)

        review = ReviewConstraints()

        table_headers_issues_table_candidate_with_issues_list = review.issues_table(default_dimensions,headers,json_table,dictionary_contraints)
        table_headers = table_headers_issues_table_candidate_with_issues_list[0]
        issues_table = table_headers_issues_table_candidate_with_issues_list[1]
        dictionary_candidate_unmatched_fields = table_headers_issues_table_candidate_with_issues_list[2]


        self.request.session['issues_table'] = issues_table
        self.request.session['table_headers'] = table_headers
        self.request.session['json_file_candidates_dictionary'] = json_table
        self.request.session['dictionary_candidate_unmatched_fields'] = dictionary_candidate_unmatched_fields
        # self.request.session['pk'] = 75

        # check if the session needs to be updated to ensure the changes are submitted
        project = Project.objects.get(pk=kwargs['pk'])
        self.request.session['pk'] = kwargs['pk']
        json_document = project.json_document
        project.json_document['issues_table'] = issues_table
        project.json_document['table_headers'] = table_headers
        project.json_document['dictionary_candidate_unmatched_fields'] = dictionary_candidate_unmatched_fields

        self.request.session['json_file'] = json_document['file_name']
        self.request.session['json_file_filename'] = json_document['file_name']
        self.request.session['json_file_headers'] = headers
        self.request.session['json_file_candidates'] = candidates
        self.request.session['json_file_candidates_list'] = candidates_list
        self.request.session['json_file_candidates_dictionary'] = candidates_dictionary


        if ((issues_table=={})):

            project.status = 'reviewed'

            project.json_document['candidates'] = candidates
            project.json_document['candidates_list'] = candidates_list
            project.json_document['candidates_dictionary'] = candidates_dictionary
            project.last_modified = datetime.now()
            project.save()

            if request.POST['Option'] == 'Save & Download':
                # if 'Save & Download' in request.POST.keys():
                return HttpResponseRedirect(reverse('viewdownload', kwargs={'pk': kwargs['pk']}))

            # with BytesIO() as b:
            #     df = pd.json_normalize(json_document['candidates'], max_level=5)
            #     # Use the StringIO object as the filehandle.
            #     writer = pd.ExcelWriter(b, engine='xlsxwriter')
            #     df.to_excel(writer, sheet_name='Sheet1')
            #     writer.save()
            #     # Set up the Http response.
            #     filename = 'django_simple.xlsx'
            #     response = HttpResponse(
            #         b.getvalue(),
            #         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            #     )
            #     response['Content-Disposition'] = 'attachment; filename=%s' % filename
            #     return response



            return super().post(request, *args, **kwargs)
        else:
            # get back on the review as errors still exist
            # create a variable to be used in the review returned after the submission
            project.json_document['candidates'] = candidates
            project.json_document['candidates_list'] = candidates_list
            project.json_document['candidates_dictionary'] = candidates_dictionary
            project.last_modified = datetime.now()
            project.save()


            self.request.session['message'] = 'please review again as issues still exist'
            # self.get_context_data(**kwargs)
            # return super().post(request, *args, **kwargs)
            if request.POST['Option'] == 'Save & Download':
                # if 'Save & Download' in request.POST.keys():
                return HttpResponseRedirect(reverse('viewdownload', kwargs={'pk': kwargs['pk']}))

            return HttpResponseRedirect(reverse('project_review_view', kwargs={'pk':kwargs['pk']}))

@method_decorator(login_required, name='dispatch')
class ProjectSubmitView(TemplateView):
    model =Project
    template_name = 'submit.html'
    context_object_name = 'submitproject'
    fields ='__all__'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['project_name'] = Project.objects.get(pk=kwargs['pk']).project_name
        return context

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        # there is no change to the json file required as that step is performed in the reviewed view

        project = Project.objects.get(pk=kwargs['pk'])
        json_document = project.json_document

        if 'issues_table' in json_document.keys():
            issues_table = json_document['issues_table']
            if issues_table != {}:
                # return a warning as there are still issue in this submission (a bag or an old implementation)
                pass

        project.status ='final'
        project.save()


        # create email for the confirmation of submission
        current_site = get_current_site(request)

        mail_subject = 'Project: ' + project.project_name + ' pk = ' + str(project.pk) + ' has been submitted'
        message = render_to_string('grouping/project_submitted.html', {
            # 'user': user.email,
            'domain': current_site.domain,
            # 'uid': urlsafe_base64_encode(force_bytes(user.pk)).decode(),
        })
        # to_email = form.cleaned_data.get('email')
        to_email = 'groupingproject.adm@gmail.com'
        email = EmailMessage(
            mail_subject, message, to=[to_email]
        )
        email.send()

        return self.get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectAnalyticsView(DetailView):
    model = Project
    template_name = 'analytics_modified.html'
    context_object_name = 'analyticsproject'
    fields ='__all__'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        calculations = Calculations()
        gender_type_value = calculations.gender_distribution(json_file=kwargs['object'].json_document)
        label_by_gender = gender_type_value[0]
        value_by_gender = gender_type_value[1]
        candidate_distribution_by_country = calculations.candidate_distribution_by_country(json_file=kwargs['object'].json_document)
        max_value = 0
        for nation in candidate_distribution_by_country:
            if nation['value'] > max_value:
                max_value = nation['value']
        top_country_by_gender = calculations.top_country_by_gender(json_file=kwargs['object'].json_document)
        gmat_distribution = calculations.gmat_distribution(json_file=kwargs['object'].json_document)
        age_distribution = calculations.age_distribution(json_file=kwargs['object'].json_document)
        gmat_age_gender =  calculations.gmat_age_gender_distribution(json_file=kwargs['object'].json_document)
        # substituted with the calculations
        # nationality = nationality_template
        # context['data'] = json.dumps(nationality).replace('"name"', 'name').replace('"value"', 'value').replace('"', '')
        # context['data'] = str(json.dumps(nationality).replace('"name"', 'name').replace('"value"', 'value'))
        # context['data'] = nationality
        total_list_dictionary_company = calculations.top_company_category(json_file=kwargs['object'].json_document)
        total_number_of_companies = total_list_dictionary_company[0]
        list_of_companies = total_list_dictionary_company[1]
        dictionary_companies = total_list_dictionary_company[2]

        context['total_number_of_companies'] = total_number_of_companies
        context['list_of_companies'] = list_of_companies
        context['dictionary_companies'] = dictionary_companies


        context['data'] = candidate_distribution_by_country
        context['label_by_gender'] = label_by_gender
        context['value_by_gender'] = value_by_gender
        context['max_value'] = max_value
        context['countries'] = top_country_by_gender[0]
        context['male'] = top_country_by_gender[1]
        context['female'] = top_country_by_gender[2]
        context['others'] = top_country_by_gender[3]
        context['gmat_range_list'] = gmat_distribution[0]
        context['gmat_range_list_male'] = gmat_distribution[1]
        context['gmat_range_list_female'] = gmat_distribution[2]
        context['gmat_range_list_others'] = gmat_distribution[3]
        context['gmat_range_list_total'] = gmat_distribution[4]

        context['age_range_list'] = age_distribution[0]
        context['age_range_list_male'] = age_distribution[1]
        context['age_range_list_female'] = age_distribution[2]
        context['age_range_list_others'] = age_distribution[3]
        context['age_range_list_total'] = age_distribution[4]

        context['gmat_age_gender_list'] = gmat_age_gender[0]
        context['gmat_age_gender_list_male'] = gmat_age_gender[1]
        context['gmat_age_gender_list_female'] = gmat_age_gender[2]
        context['gmat_age_gender_list_others'] = gmat_age_gender[3]
        context['gmat_age_gender_list_total'] = gmat_age_gender[4]


        context['project_list'] = Project.objects.all()

        return context

    def get(self, request, *args, **kwargs):

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        return super().get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectSettingsView(TemplateView):
    template_name = 'settings.html'
    context_object_name = 'settingsproject'
    fields ='__all__'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['mandatory_dimensions'] = mandatory_dimensions
        context['nationality'] = nationality_template
        context['dictionary_contraints'] = dictionary_contraints
        context['job_function'] = job_function
        context['professional_category'] = professional_category
        context['company'] = company
        context['job_title'] = job_title

        return context


    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectConstraintsView(TemplateView):
    template_name = 'constraints.html'
    context_object_name = 'constraintsproject'
    fields ='__all__'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['nationality_list'] = Nationality.objects.all()
        context['job_title'] = JobTitle.objects.all()
        context['company'] = Company.objects.all()
        context['professional_category'] = ProfessionalCategory.objects.all()
        context['job_function'] = JobFunction.objects.all()
        context['table_fields_mandatory'] = TableFieldsMandatory.objects.all() # contains the fields for the file excel that are mandatory this will be used to be able to change the headers of the excel
        return context


    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def post(self,request, *args, **kwargs):
        table_name = self.request.POST['submit']
        value = self.request.POST[table_name]
        model_class = get_model_class(table_name)
        model_class.objects.get_or_create(value=value)

        return self.get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectAnalyticsSelectView(ListView):
    model =Project
    template_name = 'analytics_select.html'
    context_object_name = 'analyticsselectproject'
    fields ='__all__'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        if request.POST['select_project']:
            id = int(request.POST['select_project'])
            return HttpResponseRedirect(reverse('project_analytics_view', kwargs={'pk':id}))
        else:
            return super().get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectUploadConstraintsView(TemplateView):
    model = Project
    template_name = 'update_constraints_table.html'
    context_object_name = 'project_upload_constraints_from_file'
    fields = '__all__'

    # paginate_by = 10

    def get(self, request, *args, **kwargs):

        if 'from_post' in kwargs.keys():
            if kwargs['from_post'] == 0:
                kwargs['json_file_headers'] = 0
                kwargs['json_file_candidates_dictionary'] = 0
                kwargs['from_post'] = 0
        else:
            kwargs['from_post'] = 0

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # context['json_file'] =self.request.session['json_file']
        # context['json_file_filename'] =  self.request.session['json_file_filename']
        # context['json_file_headers'] = self.request.session['json_file_headers']
        # context['json_file_candidates'] = self.request.session['json_file_candidates']
        # context['json_file_candidates_list'] = self.request.session['json_file_candidates_list']
        # context['json_file_candidates_dictionary'] = self.request.session['json_file_candidates_dictionary']

        return context

    def post(self, request, *args, **kwargs):

        if (len(request.FILES)!=0) :
            select_table = request.POST['select_table']
            excel_file = request.FILES["input_file"]
            json_file = excel_to_json(excel_file)

            # self.request.session['json_file'] = json_file
            # self.request.session['json_file_filename'] = json_file['file_name']
            # self.request.session['json_file_headers'] = json_file['headers']
            # self.request.session['json_file_candidates'] = json_file['candidates']
            # self.request.session['json_file_candidates_list'] = json_file['candidates_list']
            # self.request.session['json_file_candidates_dictionary'] = json_file['candidates_dictionary']
            #
            # kwargs['json_file'] =self.request.session['json_file']
            # kwargs['json_file_filename'] =self.request.session['json_file_filename']
            # kwargs['json_file_headers'] =self.request.session['json_file_headers']
            # kwargs['json_file_candidates'] =self.request.session['json_file_candidates']
            # kwargs['json_file_candidates_list'] =self.request.session['json_file_candidates_list']
            # kwargs['json_file_candidates_dictionary']=self.request.session['json_file_candidates_dictionary']

            # kwargs['json_file'] =json_file
            # kwargs['json_file_filename'] =json_file['file_name']
            # kwargs['json_file_headers'] =json_file['headers']
            # kwargs['json_file_candidates'] =json_file['candidates']

            kwargs['json_file_headers'] =json_file['headers']
            kwargs['json_file_candidates_dictionary']=json_file['candidates_dictionary']

            kwargs['from_post'] = 1
            kwargs['selected_table'] = select_table

            # get the model and query for the elements

            model_class = get_model_class(select_table)
            # table = model_class(select_table) create an object
            table_object = model_class.objects.all()
            kwargs['table_object'] = table_object
            return self.get(request, *args, **kwargs)
        else:
            # test the post and see whether it is a submission or a review
            kwargs['json_file_headers'] = 0
            kwargs['json_file_candidates_dictionary'] = 0
            kwargs['from_post'] = 1
            return self.get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectUploadedConstraintsView(TemplateView):
    model = Project
    template_name = 'updated_constraints_table.html'
    context_object_name = 'project_uploaded_constraints_from_file'
    fields = '__all__'


    # paginate_by = 10

    def get(self, request, *args, **kwargs):
        if 'from_post' in kwargs.keys():
            if kwargs['from_post'] == 0:
                kwargs['json_file_headers'] = 0
                kwargs['json_file_candidates_dictionary'] = 0
                kwargs['from_post'] = 0
        else:
            kwargs['from_post'] = 0
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    def post(self, request, *args, **kwargs):
        # value to be saved in the tables
        posted_items_tuple_list = list(request.POST.items())

        # the post contains the tuple for the key, value from the table posted for verification
        # we need to transform the post_items in a json format, however the post return a flat view for the items
        headers_json_table = table_post_to_json(posted_items_tuple_list)
        json_table = headers_json_table[1]
        headers = headers_json_table[0]
        table_selected = self.request.POST['table_selected']

        model_class = get_model_class(table_selected)



        output_uploaded = []
        output_not_uploaded = []
        for header in headers:
            for candidate in json_table:
                value = json_table[candidate][header]
                output = model_class.objects.get_or_create(value=value)
                if output[1] == False:
                    output_not_uploaded.append(value)
                elif output[1] == True:
                    output_uploaded.append(value)

        table_object = model_class.objects.all()
        kwargs['table_object'] = table_object
        kwargs['output_not_uploaded'] = output_not_uploaded
        kwargs['output_uploaded'] = output_uploaded
        kwargs['selected_table'] = table_selected
        kwargs['from_post'] = 1
        return super().get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectDeleteView(UpdateView):
    model = Project
    context_object_name = 'deleteview'
    ordering = ['-last_modified']
    fields = '__all__'
    # ordering = ['-date_elaboration']
    # paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        # check if it is confirm or cancelled. in both cases we return to the summary

        if 'confirm' in request.POST.keys():
            # delete the project
            project = Project.objects.get(pk=kwargs['pk'])
            project.delete()
            project.save()
            pass
        elif 'cancel' in request.POST.keys():
            # do nothing
            pass

        return HttpResponseRedirect(reverse('viewlist'))


@method_decorator(login_required, name='dispatch')
class ProjectCopyView(UpdateView):
    model = Project
    context_object_name = 'viewcopy'
    ordering = ['-last_modified']
    fields = '__all__'
    # ordering = ['-date_elaboration']
    # paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context

    def get(self, request, *args, **kwargs):
        project = Project.objects.get(pk=kwargs['pk'])

        newProject = Project(
            user=project.user,
            owner = project.owner,
            status = "created",
            project_name = project.project_name + "_copy",
            input_documents=project.input_documents,
            json_document=project.json_document
        )

        # if project.status = 'created': this means that the json_document in the original file doesnt exist

        newProject.save()
        return HttpResponseRedirect(reverse('viewlist'))
        # return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):



        # check if it is confirm or cancelled. in both cases we return to the summary

        # if 'confirm' in request.POST.keys():
        #     # delete the project
        #     project = Project.objects.get(pk=kwargs['pk'])
        #     project.delete()
        #     project.save()
        #     pass
        # elif 'cancel' in request.POST.keys():
        #     # do nothing
        #     pass

        return HttpResponseRedirect(reverse('viewlist'))

@method_decorator(login_required, name='dispatch')
class ProjectDownloadView(DetailView):
    model = Project
    context_object_name = 'downloadview'
    # ordering = ['-last_modified']
    fields = '__all__'
    # ordering = ['-date_elaboration']
    # paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context

    def get(self, request, *args, **kwargs):

        project = Project.objects.get(pk=kwargs['pk'])
        json_document = project.json_document
        # posted_items_tuple_list = request.session['posted_items_tuple_list']
        #
        # self.object = self.get_object()
        # headers_json_table = table_post_to_json(posted_items_tuple_list)
        # json_table = headers_json_table[1]
        # headers = headers_json_table[0]
        # candidates = headers_json_table[2]
        # candidates_list = headers_json_table[3]
        # candidates_dictionary = headers_json_table[4]
        # # jsondata  candidates_dictionary are the same (to be reivewed)
        #
        # review = ReviewConstraints()
        #
        # table_headers_issues_table_candidate_with_issues_list = review.issues_table(default_dimensions, headers,
        #                                                                             json_table, dictionary_contraints)
        # table_headers = table_headers_issues_table_candidate_with_issues_list[0]
        # issues_table = table_headers_issues_table_candidate_with_issues_list[1]
        # dictionary_candidate_unmatched_fields = table_headers_issues_table_candidate_with_issues_list[2]
        #
        # project = Project.objects.get(pk=kwargs['pk'])
        #
        # project.json_document['issues_table'] = issues_table
        # project.json_document['table_headers'] = table_headers
        # project.json_document['dictionary_candidate_unmatched_fields'] = dictionary_candidate_unmatched_fields
        #
        #
        # project.json_document['candidates'] = candidates
        # project.json_document['candidates_list'] = candidates_list
        # project.json_document['candidates_dictionary'] = candidates_dictionary
        # project.last_modified = datetime.now()
        # project.save()


        with BytesIO() as b:


            df = pd.json_normalize(json_document['candidates'], max_level=5)
            # Use the StringIO object as the filehandle.
            writer = pd.ExcelWriter(b, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1')
            writer.save()
            # Set up the Http response.
            filename = project.json_document['file_name'].split('.')[0] + '_' + str(datetime.now().date()) + '.xlsx'
            response = HttpResponse(
                b.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            response['Content-Disposition'] = 'attachment; filename=%s' % filename
            return response


        # return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        return super().get(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ProjectFinalView(DetailView):
    model = Project
    context_object_name = 'finalview'
    # ordering = ['-last_modified']
    fields = '__all__'
    # ordering = ['-date_elaboration']
    # paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        json_document = context['object'].json_document
        context['get_post'] = self.request.session['get_post']
        context['json_file'] = json_document
        context['json_file_filename'] = json_document['file_name']
        context['json_file_headers'] = json_document['headers']
        context['json_file_candidates'] = json_document['candidates']
        context['json_file_candidates_list'] = json_document['candidates_list']
        context['json_file_candidates_dictionary'] = json_document['candidates_dictionary']

        return context

    def get(self, request, *args, **kwargs):
        self.request.session['get_post'] = 'get'
        # if request.GET['Option'] == 'Save & Download':
        #     # if 'Save & Download' in request.POST.keys():
        #     return HttpResponseRedirect(reverse('viewdownload', kwargs={'pk': kwargs['pk']}))
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.request.session['get_post'] = 'post'
        return HttpResponseRedirect(reverse('viewdownload', kwargs={'pk': kwargs['pk']}))
        # return super().get(request, *args, **kwargs)

def excel_to_json(file):
    # convert excel file to json
    # with no specified number of columns on the excel
    saved_entries = []
    not_saved_entries = []
    wb = openpyxl.load_workbook(file)

    worksheet = wb[wb.sheetnames[0]]  # take the first sheet as the one with the data

    excel_data = list()
    excel_data_no_headers = list()

    # iterating over the rows and
    # getting value from each cell in row

    # table_name = model
    # model_class = getattr(sys.modules[__name__], table_name)
    i = 0  # i represent the rows

    headers_file = []
    data_dict = {}
    data_list = []
    data_dict_rows = {}
    list_dict_rows = []
    list_list_rows = []
    for row in worksheet.iter_rows():
        row_data = list()
        j = 0
        for cell in row:
            if cell.value:
                row_data.append(str(cell.value))
                if i != 0:
                    data_dict[headers_file[j]] = str(cell.value)
                    data_list.append(str(cell.value))
            else:
                row_data.append(None)
                if i != 0:
                    data_dict[headers_file[j]] = None
                    data_list.append(None)

            j = j + 1
        if i != 0:
            data_dict_rows['candidate_' + str(i)] = data_dict

            list_dict_rows.append(data_dict)

            list_list_rows.append(data_list)

        if i == 0:
            headers_file = row_data

        else:
            try:
                excel_data_no_headers.append(row_data)
                # model_class(**data_dict).save()save
                print('saved question number:%s' % i)
                saved_entries.append(i)
            except:
                print('not saved')
                not_saved_entries.append(i)
                pass
        data_dict = {}
        data_list = []
        excel_data.append(row_data)
        i = i + 1
        json_file = {
            'file_name': file.name,
            'headers':headers_file,
            'candidates': list_dict_rows,
            'candidates_list': list_list_rows,
            'candidates_dictionary': data_dict_rows,

        }
    return json_file

def table_post_to_json(posted_items_tuple_list):
    headers = []
    items = []
    for tuple in posted_items_tuple_list:  # element is a tuple
        # if "header" in tuple[0]:
        if tuple[0].startswith("header"):
            # headers.append(element[7:(len(element))])
            headers.append(tuple[1])
            pass
        # elif "item" in tuple[0]:
        elif tuple[0].startswith("item"):
            items.append(tuple[1])
            pass
        else:
            pass
    jsondata = {}
    jsondata_candidate = {}
    candidates = []
    candidates_list = []
    list_for_candidates = []
    num_of_rows = int(len(items) / len(headers))
    index = 0
    for row_num in range(num_of_rows):

        for header in headers:
            jsondata_candidate[header] = items[index]
            list_for_candidates.append(items[index])
            index = index + 1

            pass
        jsondata["candidate_" + str(row_num + 1)] = jsondata_candidate
        candidates.append(jsondata_candidate)
        candidates_list.append(list_for_candidates)
        jsondata_candidate = {}
        list_for_candidates = []


    candidates_dictionary= jsondata
    return [headers,jsondata,candidates,candidates_list,candidates_dictionary]
    # return [headers,jsondata]

def json_to_excel(json_document):
    with BytesIO() as b:
        df = pd.json_normalize(json_document['candidates'], max_level=5)
        # Use the StringIO object as the filehandle.
        writer = pd.ExcelWriter(b, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1')
        writer.save()
        # Set up the Http response.
        filename = 'django_simple.xlsx'
        response = HttpResponse(
            b.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response
    # # df = pd.json_normalize(json_document['candidates'], max_level=5)
    # # df.to_excel('test.xlsx', index=False)
    # # https://idiomaticprogrammers.com/post/how-to-download-pandas-dataframe-as-excel-or-csv-in-django-/
    # # https://stackoverflow.com/questions/63584745/django-pandas-create-excel-file-and-serve-as-download
    # output = BytesIO()
    #
    # df = pd.json_normalize(json_document['candidates'], max_level=5)
    #
    # writer = pd.ExcelWriter(output, engine='xlsxwriter')
    # df.to_excel(writer, sheet_name='Sheet1')
    # writer.save()
    #
    # output.seek(0)
    # # workbook = output.getvalue()
    #
    # output.getvalue()
    # content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    #
    # # response = HttpResponse(output, content_type=content_type)
    # # response['Content-Disposition'] = f'attachment; filename=pippo.xlsx'
    #
    # return output



nationality_template =[{
                    'name': 'Afghanistan',
                    'value': 58397.812
                }, {
                    'name': 'Angola',
                    'value': 19549.124
                }, {
                    'name': 'Albania',
                    'value': 3150.143
                }, {
                    'name': 'United Arab Emirates',
                    'value': 8441.537
                }, {
                    'name': 'Argentina',
                    'value': 40374.224
                }, {
                    'name': 'Armenia',
                    'value': 2963.496
                }, {
                    'name': 'French Southern and Antarctic Lands',
                    'value': 268.065
                }, {
                    'name': 'Australia',
                    'value': 522404.488
                }, {
                    'name': 'Austria',
                    'value': 8401.924
                }, {
                    'name': 'Azerbaijan',
                    'value': 9094.718
                }, {
                    'name': 'Burundi',
                    'value': 9232.753
                }, {
                    'name': 'Belgium',
                    'value': 10941.288
                }, {
                    'name': 'Benin',
                    'value': 9509.798
                }, {
                    'name': 'Burkina Faso',
                    'value': 15540.284
                }, {
                    'name': 'Bangladesh',
                    'value': 151125.475
                }, {
                    'name': 'Bulgaria',
                    'value': 7389.175
                }, {
                    'name': 'The Bahamas',
                    'value': 66402.316
                }, {
                    'name': 'Bosnia and Herzegovina',
                    'value': 3845.929
                }, {
                    'name': 'Belarus',
                    'value': 9491.07
                }, {
                    'name': 'Belize',
                    'value': 308.595
                }, {
                    'name': 'Bermuda',
                    'value': 64.951
                }, {
                    'name': 'Bolivia',
                    'value': 716.939
                }, {
                    'name': 'Brazil',
                    'value': 195210.154
                }, {
                    'name': 'Brunei',
                    'value': 27.223
                }, {
                    'name': 'Bhutan',
                    'value': 716.939
                }, {
                    'name': 'Botswana',
                    'value': 1969.341
                }, {
                    'name': 'Central African Republic',
                    'value': 4349.921
                }, {
                    'name': 'Canada',
                    'value': 994127.24
                }, {
                    'name': 'Switzerland',
                    'value': 7830.534
                }, {
                    'name': 'Chile',
                    'value': 17150.76
                }, {
                    'name': 'China',
                    'value': 1359821.465
                }, {
                    'name': 'Ivory Coast',
                    'value': 60508.978
                }, {
                    'name': 'Cameroon',
                    'value': 20624.343
                }, {
                    'name': 'Democratic Republic of the Congo',
                    'value': 62191.161
                }, {
                    'name': 'Republic of the Congo',
                    'value': 3573.024
                }, {
                    'name': 'Colombia',
                    'value': 46444.798
                }, {
                    'name': 'Costa Rica',
                    'value': 4669.685
                }, {
                    'name': 'Cuba',
                    'value': 11281.768
                }, {
                    'name': 'Northern Cyprus',
                    'value': 1.468
                }, {
                    'name': 'Cyprus',
                    'value': 1103.685
                }, {
                    'name': 'Czech Republic',
                    'value': 10553.701
                }, {
                    'name': 'Germany',
                    'value': 83017.404
                }, {
                    'name': 'Djibouti',
                    'value': 834.036
                }, {
                    'name': 'Denmark',
                    'value': 5550.959
                }, {
                    'name': 'Dominican Republic',
                    'value': 10016.797
                }, {
                    'name': 'Algeria',
                    'value': 37062.82
                }, {
                    'name': 'Ecuador',
                    'value': 15001.072
                }, {
                    'name': 'Egypt',
                    'value': 78075.705
                }, {
                    'name': 'Eritrea',
                    'value': 5741.159
                }, {
                    'name': 'Spain',
                    'value': 46182.038
                }, {
                    'name': 'Estonia',
                    'value': 1298.533
                }, {
                    'name': 'Ethiopia',
                    'value': 87095.281
                }, {
                    'name': 'Finland',
                    'value': 5367.693
                }, {
                    'name': 'Fiji',
                    'value': 860.559
                }, {
                    'name': 'Falkland Islands',
                    'value': 49.581
                }, {
                    'name': 'France',
                    'value': 63230.866
                }, {
                    'name': 'Gabon',
                    'value': 1556.222
                }, {
                    'name': 'United Kingdom',
                    'value': 62066.35
                }, {
                    'name': 'Georgia',
                    'value': 4388.674
                }, {
                    'name': 'Ghana',
                    'value': 24262.901
                }, {
                    'name': 'Guinea',
                    'value': 10876.033
                }, {
                    'name': 'Gambia',
                    'value': 1680.64
                }, {
                    'name': 'Guinea Bissau',
                    'value': 10876.033
                }, {
                    'name': 'Equatorial Guinea',
                    'value': 696.167
                }, {
                    'name': 'Greece',
                    'value': 11109.999
                }, {
                    'name': 'Greenland',
                    'value': 56.546
                }, {
                    'name': 'Guatemala',
                    'value': 14341.576
                }, {
                    'name': 'French Guiana',
                    'value': 231.169
                }, {
                    'name': 'Guyana',
                    'value': 786.126
                }, {
                    'name': 'Honduras',
                    'value': 7621.204
                }, {
                    'name': 'Croatia',
                    'value': 4338.027
                }, {
                    'name': 'Haiti',
                    'value': 9896.4
                }, {
                    'name': 'Hungary',
                    'value': 10014.633
                }, {
                    'name': 'Indonesia',
                    'value': 240676.485
                }, {
                    'name': 'India',
                    'value': 1205624.648
                }, {
                    'name': 'Ireland',
                    'value': 4467.561
                }, {
                    'name': 'Iran',
                    'value': 240676.485
                }, {
                    'name': 'Iraq',
                    'value': 30962.38
                }, {
                    'name': 'Iceland',
                    'value': 318.042
                }, {
                    'name': 'Israel',
                    'value': 7420.368
                }, {
                    'name': 'Italy',
                    'value': 60508.978
                }, {
                    'name': 'Jamaica',
                    'value': 2741.485
                }, {
                    'name': 'Jordan',
                    'value': 6454.554
                }, {
                    'name': 'Japan',
                    'value': 127352.833
                }, {
                    'name': 'Kazakhstan',
                    'value': 15921.127
                }, {
                    'name': 'Kenya',
                    'value': 40909.194
                }, {
                    'name': 'Kyrgyzstan',
                    'value': 5334.223
                }, {
                    'name': 'Cambodia',
                    'value': 14364.931
                }, {
                    'name': 'South Korea',
                    'value': 51452.352
                }, {
                    'name': 'Kosovo',
                    'value': 97.743
                }, {
                    'name': 'Kuwait',
                    'value': 2991.58
                }, {
                    'name': 'Laos',
                    'value': 6395.713
                }, {
                    'name': 'Lebanon',
                    'value': 4341.092
                }, {
                    'name': 'Liberia',
                    'value': 3957.99
                }, {
                    'name': 'Libya',
                    'value': 6040.612
                }, {
                    'name': 'Sri Lanka',
                    'value': 20758.779
                }, {
                    'name': 'Lesotho',
                    'value': 2008.921
                }, {
                    'name': 'Lithuania',
                    'value': 3068.457
                }, {
                    'name': 'Luxembourg',
                    'value': 507.885
                }, {
                    'name': 'Latvia',
                    'value': 2090.519
                }, {
                    'name': 'Morocco',
                    'value': 31642.36
                }, {
                    'name': 'Moldova',
                    'value': 103.619
                }, {
                    'name': 'Madagascar',
                    'value': 21079.532
                }, {
                    'name': 'Mexico',
                    'value': 117886.404
                }, {
                    'name': 'Macedonia',
                    'value': 507.885
                }, {
                    'name': 'Mali',
                    'value': 13985.961
                }, {
                    'name': 'Myanmar',
                    'value': 51931.231
                }, {
                    'name': 'Montenegro',
                    'value': 620.078
                }, {
                    'name': 'Mongolia',
                    'value': 2712.738
                }, {
                    'name': 'Mozambique',
                    'value': 23967.265
                }, {
                    'name': 'Mauritania',
                    'value': 3609.42
                }, {
                    'name': 'Malawi',
                    'value': 15013.694
                }, {
                    'name': 'Malaysia',
                    'value': 28275.835
                }, {
                    'name': 'Namibia',
                    'value': 2178.967
                }, {
                    'name': 'New Caledonia',
                    'value': 246.379
                }, {
                    'name': 'Niger',
                    'value': 15893.746
                }, {
                    'name': 'Nigeria',
                    'value': 159707.78
                }, {
                    'name': 'Nicaragua',
                    'value': 5822.209
                }, {
                    'name': 'Netherlands',
                    'value': 16615.243
                }, {
                    'name': 'Norway',
                    'value': 4891.251
                }, {
                    'name': 'Nepal',
                    'value': 26846.016
                }, {
                    'name': 'New Zealand',
                    'value': 4368.136
                }, {
                    'name': 'Oman',
                    'value': 2802.768
                }, {
                    'name': 'Pakistan',
                    'value': 173149.306
                }, {
                    'name': 'Panama',
                    'value': 3678.128
                }, {
                    'name': 'Peru',
                    'value': 29262.83
                }, {
                    'name': 'Philippines',
                    'value': 93444.322
                }, {
                    'name': 'Papua New Guinea',
                    'value': 6858.945
                }, {
                    'name': 'Poland',
                    'value': 38198.754
                }, {
                    'name': 'Puerto Rico',
                    'value': 3709.671
                }, {
                    'name': 'North Korea',
                    'value': 1.468
                }, {
                    'name': 'Portugal',
                    'value': 10589.792
                }, {
                    'name': 'Paraguay',
                    'value': 6459.721
                }, {
                    'name': 'Qatar',
                    'value': 1749.713
                }, {
                    'name': 'Romania',
                    'value': 21861.476
                }, {
                    'name': 'Russia',
                    'value': 21861.476
                }, {
                    'name': 'Rwanda',
                    'value': 10836.732
                }, {
                    'name': 'Western Sahara',
                    'value': 514.648
                }, {
                    'name': 'Saudi Arabia',
                    'value': 27258.387
                }, {
                    'name': 'Sudan',
                    'value': 35652.002
                }, {
                    'name': 'South Sudan',
                    'value': 9940.929
                }, {
                    'name': 'Senegal',
                    'value': 12950.564
                }, {
                    'name': 'Solomon Islands',
                    'value': 526.447
                }, {
                    'name': 'Sierra Leone',
                    'value': 5751.976
                }, {
                    'name': 'El Salvador',
                    'value': 6218.195
                }, {
                    'name': 'Somaliland',
                    'value': 9636.173
                }, {
                    'name': 'Somalia',
                    'value': 9636.173
                }, {
                    'name': 'Republic of Serbia',
                    'value': 3573.024
                }, {
                    'name': 'Suriname',
                    'value': 524.96
                }, {
                    'name': 'Slovakia',
                    'value': 5433.437
                }, {
                    'name': 'Slovenia',
                    'value': 2054.232
                }, {
                    'name': 'Sweden',
                    'value': 9382.297
                }, {
                    'name': 'Swaziland',
                    'value': 1193.148
                }, {
                    'name': 'Syria',
                    'value': 7830.534
                }, {
                    'name': 'Chad',
                    'value': 11720.781
                }, {
                    'name': 'Togo',
                    'value': 6306.014
                }, {
                    'name': 'Thailand',
                    'value': 66402.316
                }, {
                    'name': 'Tajikistan',
                    'value': 7627.326
                }, {
                    'name': 'Turkmenistan',
                    'value': 5041.995
                }, {
                    'name': 'East Timor',
                    'value': 10016.797
                }, {
                    'name': 'Trinidad and Tobago',
                    'value': 1328.095
                }, {
                    'name': 'Tunisia',
                    'value': 10631.83
                }, {
                    'name': 'Turkey',
                    'value': 72137.546
                }, {
                    'name': 'United Republic of Tanzania',
                    'value': 44973.33
                }, {
                    'name': 'Uganda',
                    'value': 33987.213
                }, {
                    'name': 'Ukraine',
                    'value': 46050.22
                }, {
                    'name': 'Uruguay',
                    'value': 3371.982
                }, {
                    'name': 'United States of America',
                    'value': 312247.116
                }, {
                    'name': 'Uzbekistan',
                    'value': 27769.27
                }, {
                    'name': 'Venezuela',
                    'value': 236.299
                }, {
                    'name': 'Vietnam',
                    'value': 89047.397
                }, {
                    'name': 'Vanuatu',
                    'value': 236.299
                }, {
                    'name': 'West Bank',
                    'value': 13.565
                }, {
                    'name': 'Yemen',
                    'value': 22763.008
                }, {
                    'name': 'South Africa',
                    'value': 51452.352
                }, {
                    'name': 'Zambia',
                    'value': 13216.985
                }, {
                    'name': 'Zimbabwe',
                    'value': 13076.978
                }]